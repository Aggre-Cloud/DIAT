"""
Excel generation module - FIXED VERSION
  - 章/节 combined: number + title in ONE cell
  - Columns: ID / 章 / 节 / 需求原文 / <dynamic language translations>
  - Only "Requirements" sheet (no Translation/Statistics/Validation)
  - Static headers + translation-column headers are rendered in the
    *display language* (= first target language by default) so an
    English-targeted sheet shows English headers and a Chinese-targeted
    sheet shows Chinese headers — no mixed-language headers.
"""
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os
import importlib.util

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Dynamically load config module
def load_config():
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                  "007_config", "config.py")
    spec = importlib.util.spec_from_file_location("config", config_path)
    config = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(config)
    return config

_config = load_config()
EXCEL_CONFIG = _config.EXCEL_CONFIG


class ExcelGenerator:
    """Excel file generator"""

    def __init__(self):
        """Initialize Excel generator"""
        self.header_color = EXCEL_CONFIG['header_color']
        self.row_height = EXCEL_CONFIG['row_height']

    # ------------------------------------------------------------------ #
    #  Public API
    # ------------------------------------------------------------------ #

    def generate(self, requirements, output_file, target_languages=None,
                 display_lang=None):
        """
        Generate Excel file — only one sheet: "Requirements"

        Args:
            requirements:      List of requirement dicts
            output_file:       Output file path
            target_languages:  List of lang codes (e.g. ['en','zh-cn'])
            display_lang:      Language for static + column headers.  When
                              None, defaults to the first target language
                              (falls back to 'en' if no targets given).
        """
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "translator", "002_translator/translator.py"
        )
        trans_mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(trans_mod)

        target_languages = target_languages or ['en', 'zh-cn']
        if display_lang is None:
            display_lang = target_languages[0] or 'en'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = trans_mod.sheet_title(display_lang)

        self._fill_requirements_sheet(ws, requirements,
                                      target_languages,
                                      display_lang,
                                      trans_mod)

        wb.save(output_file)
        return output_file

    # ------------------------------------------------------------------ #
    #  Sheet filling
    # ------------------------------------------------------------------ #

    @staticmethod
    def _combine(number, title):
        """Combine number + title into a single cell value."""
        num = str(number).strip() if number and str(number) != 'None' else ''
        tit = str(title).strip() if title and str(title) != 'None' else ''
        parts = [num, tit]
        parts = [p for p in parts if p]
        return ' '.join(parts)

    @staticmethod
    def _clean_artifacts(text):
        """Remove PDF-extraction cid() artefacts and normalize whitespace."""
        if not text:
            return text
        text = re.sub(r'\(cid:\d+\)', '', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text

    def _fill_requirements_sheet(self, ws, requirements,
                                 target_languages, display_lang,
                                 trans_mod):
        """Fill the single Requirements sheet with dynamic language columns.

        ``display_lang`` controls the language of both the static headers
        (ID / 章 / 节 / 需求原文) and the per-language column headers.
        Resolve helper functions defensively so the generator also works
        with older translator modules that lack the new helpers.
        """
        _static_headers = getattr(trans_mod, 'static_headers',
                                  lambda dl: ["ID", "章", "节", "需求原文"])
        _lang_col_header = getattr(trans_mod, 'lang_column_header',
                                   trans_mod.lang_label)

        # ---- Build header list dynamically --------------------------------
        # Layout: ID / 章 / 节 / 需求原文 / <dynamic language columns>
        static_headers = _static_headers(display_lang)
        lang_headers = [_lang_col_header(c, display_lang)
                        for c in target_languages]
        headers = static_headers + lang_headers
        num_cols = len(headers)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color=self.header_color,
            end_color=self.header_color,
            fill_type="solid",
        )
        thin_border = Border(
            left=Side(style='thin'),   right=Side(style='thin'),
            top=Side(style='thin'),    bottom=Side(style='thin'),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[1].height = 30

        # ---- Data rows ----------------------------------------------------
        wrap = Alignment(vertical="top", wrap_text=True)
        center_top = Alignment(horizontal="center", vertical="top")

        for row_idx, req in enumerate(requirements, 2):
            req = req if isinstance(req, dict) else {}

            original = self._clean_artifacts(
                req.get("content") or req.get("original") or ""
            )

            # chapter & section are now clean from the parser — just combine
            chapter = self._combine(
                req.get("chapter_number", ""),
                req.get("chapter_title", ""),
            )
            section = self._combine(
                req.get("section_number", ""),
                req.get("section_title", ""),
            )

            # Get translations from req['translations'] dict
            translations = req.get('translations', {}) or {}
            per_lang = {}
            for lang in target_languages:
                val = self._clean_artifacts(translations.get(lang, '') or '')
                per_lang[lang] = val

            # Write static columns (ID / 章|Chapter / 节|Section / 需求原文|Source)
            ws.cell(row=row_idx, column=1, value=req.get("id", "")).alignment = center_top
            ws.cell(row=row_idx, column=2, value=chapter).alignment   = wrap
            ws.cell(row=row_idx, column=3, value=section).alignment   = wrap
            ws.cell(row=row_idx, column=4, value=original).alignment  = wrap

            # Write dynamic language columns
            longest = len(original)
            for i, lang in enumerate(target_languages):
                col_idx = 5 + i
                val = per_lang[lang]
                ws.cell(row=row_idx, column=col_idx, value=val).alignment = wrap
                if len(val) > longest:
                    longest = len(val)

            # Borders
            for c in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=c).border = thin_border

            est_lines = max(1, longest // 65)
            ws.row_dimensions[row_idx].height = min(180, max(30, est_lines * 15))

        # ---- Column widths ------------------------------------------------
        # ID / 章|Chapter / 节|Section / 需求原文|Source / <langs...>
        col_widths = [10, 32, 32, 65] + [65] * len(target_languages)
        for col, width in enumerate(col_widths[:num_cols], 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # ---- Freeze header + auto filter ----------------------------------
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{len(requirements) + 1}"
