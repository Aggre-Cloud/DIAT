"""
DIaT — main CLI entry point and pipeline orchestration.

Responsibilities
----------------
* Parse CLI args and run the interactive / non-interactive pipeline.
* Provide `ensure_dependencies()`, a self-install check that lets the agent
  (re)install missing packages from the project `requirements.txt` and
  continue without human intervention.
* Provide `write_translations_to_excel()` for the Agent translation path.

Dependency contract
-------------------
All third-party imports the pipeline needs are listed in `REQUIRED_PACKAGES`
(below).  If any import fails, the agent should call
``python -m 006_main.main --install-deps`` (or invoke
``ensure_dependencies()``) to install them and then re-run.  Manual users
install the standard way::

    pip install -r requirements.txt
"""
import sys
import os
import subprocess
from pathlib import Path
import importlib
import importlib.util
import json

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ---------------------------------------------------------------------------
# Dependency table — the single source of truth for third-party requirements.
# Keep this synchronized with the project-root ``requirements.txt``.
#   import_name  : pip package specifier (what pip installs)
# ---------------------------------------------------------------------------
REQUIRED_PACKAGES = {
    "openpyxl":        "openpyxl==3.1.2",
    "pdfplumber":      "pdfplumber==0.11.9",
    "PyPDF2":           "PyPDF2==3.0.1",
    "pypdfium2":       "pypdfium2==5.3.0",
    "googletrans":     "googletrans==4.0.0rc1",
}

# Optional extras — not required for the core pipeline, installed on demand.
OPTIONAL_PACKAGES = {
    "pysbd":           "pysbd>=0.3",
}

# Map "nice" labels for user-facing messaging.
OPTIONAL_LABELS = {
    "pysbd": "pysbd (source-language-aware sentence segmentation; hard-coded regex fallback used if absent)",
}


def _missing_imports():
    """Return a list of import names from REQUIRED_PACKAGES that are not
    currently importable in this Python process.

    A module that exists but fails to import (corrupt install, missing
    native dep, …) is treated as missing.
    """
    missing = []
    for import_name, _ in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(import_name)
        except Exception:
            missing.append(import_name)
    return missing


def ensure_dependencies(auto_install=True, interactive=False, install_optional=False):
    """Verify every REQUIRED_PACKAGES import resolves.

    On the first failure:
      * If ``auto_install`` is True, run ``pip install`` for the missing
        packages from the project ``requirements.txt`` (so the same pinned
        versions the project ships are used), then re-check.
      * If ``auto_install`` is False and ``interactive`` is True, ask the
        user whether to install now.
      * Otherwise, print a clear remediation command and return False.

    Parameters
    ----------
    auto_install : bool
        Install missing third-party packages without prompting.  The agent
        path sets this to True.
    interactive : bool
        When True and auto_install is False, prompt the user before
        installing (manual-CLI convenience).
    install_optional : bool
        Also try to import+install the OPTIONAL_PACKAGES set (e.g. pysbd).

    Returns
    -------
    bool
        True if all required imports resolve afterwards, False otherwise.
    """
    missing = _missing_imports()
    if not missing:
        return True

    project_root = Path(__file__).resolve().parent.parent
    req_file = project_root / "requirements.txt"

    # Translate import names → pip specifiers (fall back to the bare name).
    specs_to_install = [
        REQUIRED_PACKAGES.get(n, n) for n in missing
    ]
    if install_optional:
        for name, spec in OPTIONAL_PACKAGES.items():
            try:
                importlib.import_module(name)
            except Exception:
                specs_to_install.append(spec)

    def _print_remediation():
        sys.stderr.write(
            "[ERROR] 缺少必需依赖 / missing required packages:\n"
        )
        for n in missing:
            sys.stderr.write(f"        - {n}\n")
        sys.stderr.write(
            "\n  · 人工运行 / manual:\n"
            f"        pip install -r {req_file}\n"
            "  · 自动安装 / agent (non-interactive):\n"
            f"        python -m 006_main.main --install-deps\n"
            "  · 针对缺包精准安装 / pin-point:\n"
            f"        pip install {' '.join(specs_to_install)}\n"
        )

    if not auto_install and not interactive:
        _print_remediation()
        return False

    if interactive:
        resp = input(
            f"  {len(missing)} 个缺包 missing: "
            f"{', '.join(missing)} — install now? [Y/n] "
        ).strip().lower()
        if resp and resp not in ("y", "yes", ""):
            _print_remediation()
            return False

    # Resolve the interpreter running this script so we install into the
    # right environment (venv / conda / system).
    py = sys.executable
    cmd = [py, "-m", "pip", "install"] + specs_to_install
    print(f"[deps] installing: {' '.join(specs_to_install)}")
    try:
        subprocess.check_call(cmd)
    except subprocess.CalledProcessError as e:
        _print_remediation()
        sys.stderr.write(
            f"\n[ERROR] pip install 失败 (exit code {e.returncode}). "
            "请人工重试上述 pip 命令 / re-run the pip command manually.\n"
        )
        return False

    importlib.invalidate_caches()
    still_missing = _missing_imports()
    if still_missing:
        _print_remediation()
        sys.stderr.write(
            "\n[ERROR] 安装后仍缺包 — 请检查 Python 环境与网络。/"
            "still missing after install — check Python env and network.\n"
        )
        return False

    print(f"[OK] 依赖就绪 / dependencies ready — {', '.join(missing)} installed.")
    return True

def load_config():
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '007_config', 'config.py')
    spec = importlib.util.spec_from_file_location('config', config_path)
    config = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(config)
    return config

def load_modules():
    modules = {}
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    paths = {
        'pdf_extractor': '003_pdf_extractor/pdf_extractor.py',
        'text_splitter': '001_text_splitter/text_splitter.py',
        'translator': '002_translator/translator.py',
        'excel_generator': '005_excel_generator/excel_generator.py',
        'validator': '008_validator/validator.py',
    }

    for name, rel_path in paths.items():
        full_path = os.path.join(base, rel_path)
        spec = importlib.util.spec_from_file_location(name, full_path)
        modules[name] = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(modules[name])

    return modules


class RequirementItemizationSkill:
    def __init__(self, output_dir=None, json_output=False):
        self.config = load_config()
        self.modules = load_modules()
        self.output_dir = Path(output_dir) if output_dir else Path(self.config.OUTPUT_FIXED_DIR)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.json_output = json_output
        self.json_dir = self.output_dir / "json"
        if self.json_output:
            self.json_dir.mkdir(parents=True, exist_ok=True)
        print(f'[OK] DIaT initialized')
        print(f'[OK] Output directory: {self.output_dir}')
        if self.json_output:
            print(f'[OK] JSON output directory: {self.json_dir}')

    def process_pdf(self, pdf_path, translate=True, save_json=None,
                    target_languages=None, engine=None, no_input=False):
        pdf_path = Path(pdf_path)
        if not pdf_path.exists():
            print(f'[ERROR] File not found {pdf_path}')
            return None
        if pdf_path.suffix.lower() != '.pdf':
            print(f'[ERROR] Not a PDF file {pdf_path}')
            return None

        print(f"\n{'='*60}")
        print(f'Processing: {pdf_path.name}')
        print(f"{'='*60}")

        # ── 1. Extract ───────────────────────────────────────────────
        print('  [1/4] Extracting PDF text...')
        pdf_extractor = self.modules['pdf_extractor'].PDFExtractor()
        text, extraction_meta = pdf_extractor.extract_text(pdf_path)
        print(f'  Extraction: pages={extraction_meta.pages} '
              f'scanned_fallback={extraction_meta.scanned_fallback_used} '
              f'blocks_stripped={extraction_meta.page_blocks_stripped}')

        # ── 2. Split (hierarchical decomposition + sentence segment) ─
        print('  [2/4] Splitting requirements (hierarchical decomposition)...')
        splitter = self.modules['text_splitter']
        parse_result = splitter.parse_text(text)
        requirements = parse_result.items
        print(f'  Found {len(requirements)} requirements  '
              f'(top-level 章={parse_result.meta["top_level_sections"]}, '
              f'nodes with body={parse_result.meta["nodes_with_body"]})')

        if not requirements:
            print('  [WARN] No requirements found, skipping')
            return None

        # ── 2b. Body-preservation assertion ─────────────────────────
        print('  [2b/4] Validating body preservation...')
        validator = self.modules['validator']
        cfg = self.config.VALIDATION
        orphans_path = str(self.output_dir / f'{pdf_path.stem}_orphans.json')
        try:
            report = validator.assert_body_intact(
                text, requirements,
                min_word_cov=cfg.get('min_word_coverage', 0.85),
                min_char_cov=cfg.get('min_char_coverage', 0.80),
                orphans_path=orphans_path if cfg.get('write_orphans', True) else None,
            )
            print(f'  Body Preservation: coverage={report.coverage_ratio:.1%} '
                  f'orphans={len(report.orphan_lines)}')
        except validator.BodyLossError as e:
            print(f'  [FATAL] {e}')
            raise

        # Save JSON if requested
        json_data = {
            'version': '3.0',
            'metadata': parse_result.meta,
            'extraction': {
                'pages': extraction_meta.pages,
                'scanned_fallback_used': extraction_meta.scanned_fallback_used,
                'page_blocks_stripped': extraction_meta.page_blocks_stripped,
                'strategy_counts': extraction_meta.strategy_counts,
            },
            'body_coverage': {
                'ratio': report.coverage_ratio,
                'orphan_count': len(report.orphan_lines),
            },
            'requirements': requirements,
        }
        if self.json_output or save_json:
            json_path = self.json_dir / f'{pdf_path.stem}_requirements.json'
            splitter.save_json_output(json_data, str(json_path))
            print(f'  [OK] JSON saved: {json_path}')

        # ── Build processed_reqs (classifier columns removed) ───────
        processed_reqs = []
        for idx, req_data in enumerate(requirements, 1):
            processed_reqs.append({
                'id': req_data.get('id', f'REQ-{idx:04d}'),
                'chapter_number': req_data.get('chapter_number', ''),
                'chapter_title': req_data.get('chapter_title', ''),
                'section_number': req_data.get('section_number', ''),
                'section_title': req_data.get('section_title', ''),
                'hierarchy_path': req_data.get('hierarchy_path', ''),
                'content': req_data.get('content', ''),
                'page': req_data.get('page', 0),
                'tag': req_data.get('tag', ''),
                'translations': {},
                'detected_source': '',
                'is_valid': req_data.get('is_valid', True),
                'validation': req_data.get('validation', {}),
            })

        if not translate:
            # Skip translation entirely
            pass
        else:
            # ── 3a. Detect source language from a sample ────────────
            translator_module = self.modules['translator']
            translator_service = translator_module.TranslationService(
                target_languages=target_languages or ['en', 'zh-cn']
            )

            # Detect source: use longest content string as sample
            sample_text = max((r['content'] for r in processed_reqs
                               if r['content']), key=len, default='')
            detected_source = translator_service.detect_language(sample_text)
            print(f'\n  Detected source language: {detected_source}')

            # ── 3b. Ask user for target languages ─────────────────
            # (no_input / CLI path) — Requirement 3: if source matches
            # one of the targets, that column keeps the original text
            # instead of calling the translation API.
            if target_languages is None:
                if no_input:
                    target_languages = ['en', 'zh-cn']
                else:
                    target_languages, _ = self._ask_target_languages(detected_source)

            # Warn if source is among targets (will keep original, not error)
            if detected_source and detected_source in target_languages:
                name = (RequirementItemizationSkill.AVAILABLE_LANGS
                        .get(detected_source, detected_source))
                print(f'  ⚠ Source={detected_source} ({name}) is in targets — '
                      f'column "{name}" will keep original text.')

            # Update targets used by Google translator service
            translator_service.target_languages = target_languages

            # ── 3c. Ask user for translation engine ───────────────
            if engine is None:
                if no_input:
                    engine = 'google'
                else:
                    engine = self._ask_translation_engine()

            # ── 3c½. Ask for extra DO_NOT_TRANSLATE terms ────────
            # Now category-guided: returns {category: [term, ...]}.
            extra_terms: dict[str, list[str]] = {}
            if not no_input:
                extra_terms = self._ask_do_not_translate_additions()
                if extra_terms:
                    # Google path: flatten categories -> extend cache
                    flat_extras = [
                        t for terms in extra_terms.values() for t in terms
                    ]
                    translator_service.do_not_translate.extend(flat_extras)
                    print(f'  + added {len(flat_extras)} term(s) across '
                          f'{len(extra_terms)} categor(ies) to '
                          f'DO_NOT_TRANSLATE:')
                    for ckey, terms in extra_terms.items():
                        if terms:
                            print('      [{}]: {}'.format(ckey, ', '.join(terms)))

            # ── 3d. Translate ─────────────────────────────────────
            print(f'\n  [3/4] Translating (engine={engine}, '
                  f'languages={target_languages})...')

            if engine == 'agent':
                self._translate_with_agent(processed_reqs, target_languages,
                                           detected_source,
                                           extra_terms=extra_terms)
                # Save partial JSON (translations empty) for agent to read
                agent_json_path = self.output_dir / 'json' / f'{pdf_path.stem}_agent_queue.json'
                self.output_dir.mkdir(parents=True, exist_ok=True)
                (self.output_dir / 'json').mkdir(parents=True, exist_ok=True)
                import json as _json
                with open(agent_json_path, 'w', encoding='utf-8') as f:
                    # extra_do_not_translate is a categorized dict
                    # ({category: [term, ...]}); the agent must call
                    # TranslationService.merge_do_not_translate_terms() to
                    # re-inject these before translating.
                    _json.dump({
                        'source_language': detected_source,
                        'target_languages': target_languages,
                        'excel_output': '',
                        'source_is_target': (
                            detected_source in target_languages
                        ),
                        'extra_do_not_translate': extra_terms,
                        'requirements': [
                            {
                                'id': r['id'],
                                'content': r['content'],
                                'translations': {
                                    lang: (r['content']
                                           if lang == detected_source
                                           else '')
                                    for lang in target_languages
                                }
                            }
                            for r in processed_reqs
                        ]
                    }, f, ensure_ascii=False, indent=2)
                print(f'  [Agent] Queue written: {agent_json_path}')
            else:
                self._translate_with_google(processed_reqs, target_languages,
                                            translator_service,
                                            detected_source)

            # Backfill detected_source + which column is "same as source"
            # (those columns keep original text — no translation needed)
            for r in processed_reqs:
                r['detected_source'] = detected_source
                r['source_is_target'] = detected_source in target_languages

        # ── 4. Generate Excel ────────────────────────────────────────
        print('\n  [4/4] Generating Excel...')
        output_name = f'{pdf_path.stem}_requirements.xlsx'
        output_file = self.output_dir / output_name
        if output_file.exists():
            try:
                output_file.unlink()
            except PermissionError:
                output_file = self.output_dir / f'{pdf_path.stem}_requirements_new.xlsx'

        excel_target_languages = target_languages or ['en', 'zh-cn']
        excel_gen = self.modules['excel_generator'].ExcelGenerator()
        excel_gen.generate(
            processed_reqs,
            str(output_file),
            target_languages=excel_target_languages,
        )

        valid_count = sum(1 for r in processed_reqs if r.get('is_valid', True))

        print(f'\n  [OK] Completed!')
        print(f'  [OK] Output file: {output_file}')
        print(f'  [OK] Total requirements: {len(processed_reqs)}')
        if len(processed_reqs) > 0:
            print(f'  [OK] Valid requirements: {valid_count} ({valid_count/len(processed_reqs)*100:.1f}%)')
        print(f'  [OK] Body coverage: {report.coverage_ratio:.1%}')

        # ── Agent-engine hand-off ──────────────────────────────────
        if translate and engine == 'agent':
            agent_json_file = self.output_dir / 'json' / f'{pdf_path.stem}_agent_queue.json'
            print('\n' + '=' * 55)
            print('  AGENT TRANSLATION MODE')
            print('=' * 55)
            print(f'  Source language : {detected_source}')
            print(f'  Target languages: {excel_target_languages}')
            print(f'  Items to translate: {len(processed_reqs)}')
            print(f'  JSON queue : {agent_json_file}')
            print(f'  Excel file : {output_file}')
            print()
            print('  NEXT STEP — the agent must:')
            print('    1. Read the JSON queue')
            print('    2. Translate each requirement per target language')
            print('    3. Write translations back into Excel columns')
            print('=' * 55)

        return str(output_file)

    # ── Interactive prompts ─────────────────────────────────────────────

    AVAILABLE_LANGS = {
        'en':    'English',
        'zh-cn': '中文（简体）',
        'pt':    'Português',
        'es':    'Español',
        'fr':    'Français',
        'de':    'Deutsch',
        'ja':    '日本語',
        'ko':    '한국어',
    }

    @staticmethod
    def _ask_target_languages(detected_source=None):
        """Interactive prompt — ask user to pick any 2 target languages.

        Default: ['en', 'zh-cn'] (English + Simplified Chinese).
        If one of the picks matches the source language, that column keeps
        the original text (no translation API call).

        Returns (final_targets, source_in_targets) where:
          final_targets    — list of 2 language codes to translate to
          source_in_targets — True if source language is among the picks
        """
        print('\n' + '=' * 55)
        print('  目标翻译语言选择 / Target Language Selection')
        print('=' * 55)
        if detected_source and detected_source in RequirementItemizationSkill.AVAILABLE_LANGS:
            print(f'  Detected source: {detected_source} '
                  f'({RequirementItemizationSkill.AVAILABLE_LANGS[detected_source]})')
        print('\n  Choose TWO languages for the translation columns.')
        print('  Default: en + zh-cn')
        for code, name in RequirementItemizationSkill.AVAILABLE_LANGS.items():
            marker = ' ← source' if code == detected_source else ''
            print(f'    {code:8s} — {name}{marker}')
        print()

        user_input = input(
            '  Enter 2 language codes (comma-separated) '
            'or press Enter for default [en,zh-cn]: '
        ).strip().lower()

        if not user_input:
            targets = ['en', 'zh-cn']
        else:
            codes = [c.strip() for c in user_input.split(',')]
            valid = [c for c in codes
                     if c in RequirementItemizationSkill.AVAILABLE_LANGS]
            # Deduplicate while preserving order
            seen = set()
            dedup = []
            for c in valid:
                if c not in seen:
                    seen.add(c)
                    dedup.append(c)
            if len(dedup) >= 2:
                targets = dedup[:2]
            elif len(dedup) == 1:
                # give a different default second language
                fallback = ('zh-cn' if dedup[0] != 'zh-cn' else 'en')
                targets = [dedup[0], fallback]
                print(f'  Only one valid code given — adding {fallback} '
                      f'as the second language.')
            else:
                print(f'  /!\\  Invalid codes {codes!r}, '
                      f'falling back to en,zh-cn.')
                targets = ['en', 'zh-cn']

        source_in_targets = (detected_source in targets)
        if source_in_targets:
            name = (RequirementItemizationSkill.AVAILABLE_LANGS
                    .get(detected_source, detected_source))
            print(f'  ⚠ Source is {detected_source} ({name}) → '
                  f'the {name} column will show the original text.')
        else:
            skipped = False

        return targets, source_in_targets

    @staticmethod
    def _ask_translation_engine():
        """Interactive prompt — pick translation engine."""
        print('\n' + '~' * 55)
        print('  翻译引擎选择 / Translation Engine')
        print('~' * 55)
        print('    1. Google Translate API   (default — fast, external)')
        print('    2. Agent  — Claude reads JSON, translates, writes back')
        print()
        user_input = input(
            '  Enter 1 or 2 (or press Enter for default Google): '
        ).strip()

        if user_input == '2':
            return 'agent'
        return 'google'

    # Ordered list of user-fillable categories (key -> bilingual label).
    # Categories with empty seed items in config; the numbered menu is
    # built from this list.  Runtime categories (user-created) are
    # handled separately via the "new category" branch.
    _USER_FILLABLE_CATEGORIES = [
        ('person_name',       {'en': 'Person Names',                'zh': '人名'}),
        ('place_name',        {'en': 'Place Names',                 'zh': '地名'}),
        ('project_code',      {'en': 'Product / Project Codes',     'zh': '产品/项目代号'}),
        ('company_local',     {'en': 'Company (this document)',     'zh': '公司名（本文档）'}),
        ('regulatory_body',   {'en': 'Regulatory Bodies',           'zh': '监管机构'}),
        ('legal_reference',   {'en': 'Legal / Doc References',      'zh': '法规/文档编号'}),
        ('industry_term',     {'en': 'Industry-Specific Terms',     'zh': '行业专有术语'}),
        ('role_title',        {'en': 'Roles / Responsibilities',    'zh': '岗位/职责'}),
    ]

    @staticmethod
    def _load_proper_noun_config():
        """Lazy-load config module, return module object."""
        import importlib.util as _ilu
        _cfg_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            '007_config', 'config.py')
        _spec = _ilu.spec_from_file_location('config_cfg', _cfg_path)
        _cfg = _ilu.module_from_spec(_spec)
        _spec.loader.exec_module(_cfg)
        return _cfg

    @staticmethod
    def _format_category_line(cat_key, label, items, max_show=8):
        """Render one summary line: ``[标签] 词1, 词2, … (N)``."""
        n = len(items)
        if n == 0:
            detail = '(empty / 空)'
        else:
            shown = items[:max_show]
            tail = f' … (+{n - max_show} more)' if n > max_show else ''
            detail = ', '.join(shown) + tail
        return '    [{}] / [{}]  {}  ({})'.format(
            label['zh'], label['en'], detail, n)

    @staticmethod
    def _ask_do_not_translate_additions() -> dict[str, list[str]]:
        """Category-guided prompt: user fills terms category-by-category.

        Returns
        -------
        dict[str, list[str]]
            ``{category_key: [term, …]}``  — the user-supplied extras,
            grouped by category.  Empty dict = nothing added.
        """
        cfg = RequirementItemizationSkill._load_proper_noun_config()
        cats = cfg.get_do_not_translate_categories()

        # Mutable per-session overlay (starts empty; we only collect
        # user additions, not the built-in seed).
        extra_terms: dict[str, list[str]] = {}

        # Seed counts/labels come from config; runtime additions show
        # up only inside this session's `extra_terms`.
        def _seed_items(key):
            return cats.get(key, {}).get('items', [])

        print('\n' + '~' * 60)
        print('  专有名词保护 / Proper-Noun Protection')
        print('~' * 60)
        print('  翻译时下列类别的术语将保持原文不译。')
        print('  The following categories are auto-protected:\n')

        # ── 1) Show seed categories (already protected, read-only) ──
        seed_shown = False
        for ckey, cat in cats.items():
            items = cat.get('items', [])
            if not items:
                continue
            seed_shown = True
            label = cat.get('label', {'en': ckey, 'zh': ckey})
            print(RequirementItemizationSkill._format_category_line(
                ckey, label, items))
        if not seed_shown:
            print('    (no seed categories configured)')
        print()

        # ── 2) Numbered menu of user-fillable + custom categories ──
        # Re-render the menu until the user picks 0 (finish).
        # Dynamic: seeded 'user-fillable' keys that already have items
        # still appear (user may add more); runtime categories from a
        # previous iteration are listed too.
        while True:
            menu = list(RequirementItemizationSkill._USER_FILLABLE_CATEGORIES)
            # Append any runtime categories the user already created
            # this session that aren't in the static menu.
            for k in extra_terms:
                if k not in [k for k, _ in menu]:
                    lbl_v = cats.get(k, {}).get('label', {'en': k, 'zh': k})
                    menu.append((k, lbl_v))

            print('  以下类别暂无术语或需按本文档补充：')
            print('  Please fill the categories that apply to this document:\n')
            for idx, (ckey, label) in enumerate(menu, 1):
                cur = extra_terms.get(ckey, [])
                n_cur = len(cur)
                extra = '  <- already {}'.format(n_cur) if n_cur else ''
                print('    {:2d}. {} / {}{}'.format(
                    idx, label['zh'], label['en'], extra))
            print('    10. ＋ 新建类别 / New category…')
            print('     0. 完成并继续 / Done — start translation')
            print()

            raw = input(
                '  请选择类别编号 (1-10 补充, 0=完成): '
            ).strip()

            if not raw:
                continue
            if not raw.isdigit():
                print('  /!\\  请输入数字编号。\n')
                continue
            choice = int(raw)

            if choice == 0:
                break
            elif choice == 10:
                # New category branch
                zh = input('    中文类别名 (label_zh): ').strip()
                en = input('    英文类别名 (label_en): ').strip()
                if not zh and not en:
                    print('  /!\\  至少输入一个类别名，已取消。\n')
                    continue
                en = en or zh
                zh = zh or en
                # Auto key from english label (snake_case), fallback custom_N
                import re as _re
                auto_key = _re.sub(r'[^A-Za-z0-9]+', '_', en).strip('_').lower()
                if not auto_key:
                    auto_key = 'custom_' + str(len(extra_terms) + 1)
                terms_raw = input(
                    f'    [{zh}/{en}] 输入逗号分隔词 (Enter 跳过): '
                ).strip()
                terms = [t.strip() for t in terms_raw.split(',')
                         if t.strip()] if terms_raw else []
                extra_terms[auto_key] = terms
                if terms:
                    # Register the new category in config overlay (label)
                    cats.setdefault(auto_key,
                                    {'label': {'en': en, 'zh': zh},
                                     'items': []})
                    print(f'    + added {len(terms)} term(s) to '
                          f'[{zh}/{en}].\n')
                else:
                    print()
                continue

            if choice < 1 or choice > len(menu):
                print(f'  /!\\  编号越界，请输入 0-{len(menu)} 之间的数字。\n')
                continue

            ckey, label = menu[choice - 1]
            cur = extra_terms.get(ckey, [])

            head = '  -> [{}] / [{}]'.format(label['zh'], label['en'])
            if cur:
                shown = ', '.join(cur[:8])
                tail = ' ...' if len(cur) > 8 else ''
                print('{}  当前已有 ({}): {}{}'.format(
                    head, len(cur), shown, tail))
            else:
                print('{}  当前为空 / currently empty'.format(head))
            terms_raw = input(
                '    输入逗号分隔词追加 (Enter 跳过): '
            ).strip()
            if terms_raw:
                new_terms = [t.strip() for t in terms_raw.split(',')
                             if t.strip()]
                existing = set(cur)
                added = [t for t in new_terms if t not in existing]
                extra_terms.setdefault(ckey, []).extend(added)
                print(f'    + added {len(added)} term(s)。\n')
            else:
                print()
            # Loop re-renders the menu with updated counts.

        total = sum(len(v) for v in extra_terms.values())
        if total:
            print(f'  ✓ 专有名词保护配置完成。共 {len(extra_terms)} 个类别，'
                  f'{total} 个术语。\n')
        else:
            print('  ✓ 无额外专有名词补充。\n')
        return extra_terms

    def _translate_with_agent(self, processed_reqs, target_languages,
                              detected_source, extra_terms=None):
        """Agent translation path — no SDK call here.

        Translations are intentionally left blank; the orchestrating agent
        (Claude) reads the generated JSON after process_pdf finishes,
        re-injects ``extra_do_not_translate`` via
        ``TranslationService.merge_do_not_translate_terms``, and fills in
        the translations.

        Parameters
        ----------
        extra_terms : dict[str, list[str]] | None
            Categorized DO_NOT_TRANSLATE additions
            (``{category: [term, …]}``).  Embedded in the agent JSON queue.
        """
        if not extra_terms:
            extra_terms = {}
        total = sum(len(v) for v in extra_terms.values())
        for lang in target_languages:
            print(f'  [Agent] Lang {lang} — left blank for agent to fill.')
        if total:
            print(f'  [Agent] {total} extra DO_NOT_TRANSLATE term(s) across '
                  f'{len(extra_terms)} categor(ies) queued for agent '
                  f'(agent should call merge_do_not_translate_terms).')
        # Placeholder hint written into first requirement note
        if processed_reqs:
            processed_reqs[0]['note'] = (
                'TRANSLATION_PENDING source={} targets={} '
                'do_not_translate={}'.format(
                    detected_source, target_languages, extra_terms)
            )

    def _translate_with_google(self, processed_reqs, target_languages,
                               translator_service, detected_source):
        """Google translation path — one-by-one (current behaviour).

        If detected_source matches one of the target languages, that column
        is filled with the original text (no API call).
        """
        for idx, req in enumerate(processed_reqs, 1):
            if idx % 10 == 0 or idx == len(processed_reqs):
                print(f'    Processing {idx}/{len(processed_reqs)}...')
            translations = translator_service.translate_to_languages(
                req['content'])
            # Requirement 3: keep original in the source-matching column
            if detected_source and detected_source in target_languages:
                translations[detected_source] = req['content']
            req['translations'] = translations
            req['detected_source'] = translations.get('detected_source', '')

    def process_folder(self, folder_path, translate=True, target_languages=None,
                       engine=None, no_input=False):
        folder_path = Path(folder_path)
        if not folder_path.exists():
            print(f'[ERROR] Folder not found {folder_path}')
            return []
        pdf_files = list(folder_path.glob('*.pdf'))
        if not pdf_files:
            print(f'No PDF files found in: {folder_path}')
            return []

        # Interactive prompts ONCE for the whole folder
        if translate and (target_languages is None or engine is None):
            if no_input:
                if target_languages is None:
                    target_languages = ['en', 'zh-cn']
                if engine is None:
                    engine = 'google'
            else:
                # Quick source detection from first PDF
                if target_languages is None:
                    first_text, _ = self.modules['pdf_extractor'].PDFExtractor().extract_text(pdf_files[0])
                    tmp = self.modules['translator'].TranslationService()
                    detected = tmp.detect_language(first_text[:1000] if first_text else '')
                    target_languages, _ = self._ask_target_languages(detected)
                if engine is None:
                    engine = self._ask_translation_engine()

        print(f'\nFound {len(pdf_files)} PDF files  '
              f'languages={target_languages}  engine={engine}')
        print(f'Output directory: {self.output_dir}')
        print('='*60)

        results = []
        for idx, pdf_file in enumerate(pdf_files, 1):
            print(f'\n[{idx}/{len(pdf_files)}] Processing: {pdf_file.name}')
            output_file = self.process_pdf(pdf_file, translate,
                                           target_languages=target_languages,
                                           engine=engine,
                                           no_input=no_input)
            if output_file:
                results.append(output_file)
        return results


def process_file(pdf_path, output_dir=None, translate=True, json_output=False,
                 target_languages=None, engine=None, no_input=False):
    skill = RequirementItemizationSkill(output_dir=output_dir, json_output=json_output)
    return skill.process_pdf(pdf_path, translate,
                             target_languages=target_languages,
                             engine=engine,
                             no_input=no_input)

def process_folder(folder_path, output_dir=None, translate=True, json_output=False,
                   target_languages=None, engine=None, no_input=False):
    skill = RequirementItemizationSkill(output_dir=output_dir, json_output=json_output)
    return skill.process_folder(folder_path, translate,
                                target_languages=target_languages,
                                engine=engine,
                                no_input=no_input)

def main():
    import argparse
    parser = argparse.ArgumentParser(
        description='DIaT — PDF requirement extraction + translation')
    parser.add_argument('input', nargs='?', default=None,
                        help='PDF file or folder path (optional when using '
                             '--install-deps).')
    parser.add_argument('-o', '--output', default=None, help='Output directory')
    parser.add_argument('--no-translate', action='store_true',
                        help='Skip translation')
    parser.add_argument('--json', action='store_true',
                        help='Save JSON intermediate output')
    parser.add_argument('-l', '--lang', default=None,
                        help='Target languages, comma-separated (e.g. en,zh-cn). '
                             'If not set, will prompt interactively.')
    parser.add_argument('-e', '--engine', default=None,
                        choices=['google', 'agent'],
                        help='Translation engine (default: prompt or google).')
    parser.add_argument('--no-input', action='store_true',
                        help='Non-interactive mode: use en,zh-cn + Google.')
    parser.add_argument('--install-deps', action='store_true',
                        help='Install missing third-party packages from '
                             'requirements.txt, then exit.  In a non-TTY '
                             '(agent) context this runs without prompting; '
                             'in a TTY it asks for confirmation first.')
    parser.add_argument('--with-optional', action='store_true',
                        help='Together with --install-deps, also install the '
                             'optional extras (e.g. pysbd).')
    args = parser.parse_args()

    # --install-deps: self-install then exit (no PDF processing).
    if args.install_deps:
        auto = not sys.stdin.isatty()   # non-interactive in agent / pipe
        ok = ensure_dependencies(
            auto_install=auto,
            interactive=not auto,
            install_optional=args.with_optional,
        )
        sys.exit(0 if ok else 1)

    # Normal run: make sure deps are present before doing any work.
    if not args.no_input:
        # Interactive human run — ask before installing.
        if not ensure_dependencies(auto_install=False, interactive=True):
            sys.exit(1)
    else:
        # Non-interactive (agent / batch) — install automatically.
        if not ensure_dependencies(auto_install=True):
            sys.exit(1)

    if args.input is None:
        parser.error('the following arguments are required: input '
                     '(unless --install-deps is used)')

    # Parse target languages from CLI or None (will prompt)
    target_languages = None
    if args.lang:
        target_languages = [c.strip() for c in args.lang.split(',')]
    if args.no_input and target_languages is None:
        target_languages = ['en', 'zh-cn']

    skill = RequirementItemizationSkill(output_dir=args.output,
                                        json_output=args.json)
    input_path = Path(args.input)

    if input_path.is_file():
        skill.process_pdf(input_path,
                          translate=not args.no_translate,
                          target_languages=target_languages,
                          engine=args.engine,
                          no_input=args.no_input)
    elif input_path.is_dir():
        skill.process_folder(input_path,
                             translate=not args.no_translate,
                             target_languages=target_languages,
                             engine=args.engine,
                             no_input=args.no_input)
    else:
        print(f'[ERROR] Path not found {input_path}')

# =====================================================================
#  Agent helpers —— called by the agent AFTER process_pdf agent-mode
# =====================================================================

def write_translations_to_excel(excel_path, translations, output_path=None):
    """Write agent-produced translations into the Excel file.

    translations: dict of dicts  {req_id: {lang_code: translated_text, ...}}
    output_path  : optional — if None, overwrites excel_path
    """
    import openpyxl
    from openpyxl.styles import Alignment

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active if wb.sheetnames else None
    if not ws:
        raise ValueError(f'No sheet found in {excel_path}')

    # Find language columns from header row
    header_to_col = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            header_to_col[str(val)] = col

    # Reverse lookup: lang code → column number via LANG_LABELS
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from importlib.util import spec_from_file_location
    spec = spec_from_file_location('translator', os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        '002_translator', 'translator.py'))
    tm = __import__('importlib.util').util.module_from_spec(spec)
    spec.loader.exec_module(tm)

    lang_col = {}
    for code, label in tm.LANG_LABELS.items():
        if label in header_to_col:
            lang_col[code] = header_to_col[label]

    # Find ID column (first column with header "ID")
    id_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value and str(ws.cell(1, col).value).upper() == 'ID':
            id_col = col
            break
    if not id_col:
        id_col = 1  # assume first column

    wrap = Alignment(vertical='top', wrap_text=True)
    updated = 0
    for row in range(2, ws.max_row + 1):
        row_id = ws.cell(row, id_col).value
        if not row_id or row_id not in translations:
            continue
        for lang_code, text in translations[row_id].items():
            col = lang_col.get(lang_code)
            if col:
                cell = ws.cell(row, col, value=text)
                cell.alignment = wrap
                updated += 1

    out = output_path or excel_path
    wb.save(out)
    print(f'[Agent] Wrote {updated} translations to {out}')
    return out


if __name__ == "__main__":
    main()
